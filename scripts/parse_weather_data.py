#!/usr/bin/env python3
"""
Parse Weather data-Monitored.xlsx and convert to JSON format for the thermal DT system.
Each sheet represents a different test point with different measurement parameters.
"""

import json
import os
from datetime import datetime
import openpyxl

# Test point coordinates (from the field measurement setup)
TESTPOINT_COORDINATES = {
    "Testpoint-1": {"lat": 22.419000, "lng": 114.207738},
    "Testpoint-2": {"lat": 22.419548, "lng": 114.208326},
    "Testpoint-3": {"lat": 22.419937, "lng": 114.206832},
    "Testpoint-4": {"lat": 22.420221, "lng": 114.203237},
    "Testpoint-5": {"lat": 22.419147, "lng": 114.204707},
    "Testpoint-6": {"lat": 22.418473, "lng": 114.204404},
    "Testpoint-7": {"lat": 22.418608, "lng": 114.205645},
    "Testpoint-9": {"lat": 22.418964, "lng": 114.207135},
    "Testpoint-10": {"lat": 22.419745, "lng": 114.205381},
    "Testpoint-11": {"lat": 22.419745, "lng": 114.205381},
    "Testpoint-12": {"lat": 22.418964, "lng": 114.207135},
    "Testpoint-13": {"lat": 22.418964, "lng": 114.207135},
}

# Device types for each test point
DEVICE_TYPES = {
    "Testpoint-1": {"type": "HOBO MX", "sensors": ["Ta", "RH", "Tdew", "Lu"], "category": "hobo"},
    "Testpoint-2": {"type": "HOBO MX", "sensors": ["Ta", "RH", "Tdew", "Lu"], "category": "hobo"},
    "Testpoint-3": {"type": "HOBO MX", "sensors": ["Ta", "RH", "Tdew", "Lu"], "category": "hobo"},
    "Testpoint-4": {"type": "HOBO MX", "sensors": ["Ta", "RH", "Tdew", "Lu"], "category": "hobo"},
    "Testpoint-5": {"type": "HOBO MX", "sensors": ["Ta", "RH", "Tdew", "Lu"], "category": "hobo"},
    "Testpoint-6": {"type": "HOBO MX", "sensors": ["Ta", "RH", "Tdew", "Lu"], "category": "hobo"},
    "Testpoint-7": {"type": "HOBO MX", "sensors": ["Ta", "RH", "Tdew", "Lu"], "category": "hobo"},
    "Testpoint-8": {"type": "HOBO MX", "sensors": ["Ta", "RH", "Tdew", "Lu"], "category": "hobo"},
    "Testpoint-9": {"type": "Weather Station", "sensors": ["Ta", "RH", "Tg", "Va", "Da", "Rsol", "PM2.5", "PM10"], "category": "weather_station"},
    "Testpoint-10": {"type": "Weather Station", "sensors": ["Ta", "RH", "Tg", "Va", "Da", "Rsol", "PM2.5", "PM10"], "category": "weather_station"},
    "Testpoint-11": {"type": "Thermocouple", "sensors": ["LST"], "category": "thermocouple"},
    "Testpoint-12": {"type": "Thermocouple", "sensors": ["LST"], "category": "thermocouple"},
    "Testpoint-13": {"type": "Radiation Tracker", "sensors": ["DHI", "DNI", "GHI"], "category": "radiation"},
}

# Sheet name to parameter mapping
SHEET_TO_PARAM = {
    "HOBO_Temp": "air_temperature",
    "HOBO_RH": "relative_humidity",
    "HOBO_Light": "light_intensity",
    "HOBO_Dew_point": "dew_point",
    "Thermocouple_Temp": "surface_temperature",
    "GlobE_temp": "globe_temperature",
    "Wind_direction": "wind_direction",
    "Wind_speed": "wind_speed",
    "Solar_radiation": "solar_radiation",
    "Air_temp": "air_temperature_ws",
    "RH": "relative_humidity_ws",
    "PM10": "pm10",
    "PM25": "pm25",
    "Pressure": "atmospheric_pressure",
    "Radiation": "radiation_components",
}

def parse_excel_file(filepath):
    """Parse the Excel file and extract data from all sheets."""
    print(f"Loading Excel file: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    all_data = {}
    testpoint_data = {}  # Organized by testpoint
    
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        sheet = wb[sheet_name]
        
        # Get header row (first row)
        headers = []
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            headers.append(str(val) if val else f"col_{col}")
        
        # Parse data rows
        sheet_data = []
        for row in range(2, sheet.max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                val = sheet.cell(row=row, column=col).value
                if val is not None:
                    # Handle datetime
                    if isinstance(val, datetime):
                        row_data[header] = val.isoformat()
                    else:
                        row_data[header] = val
            if row_data:  # Only add non-empty rows
                sheet_data.append(row_data)
        
        all_data[sheet_name] = {
            "headers": headers,
            "data": sheet_data,
            "row_count": len(sheet_data)
        }
        
        # Extract testpoint-specific data
        param_name = SHEET_TO_PARAM.get(sheet_name, sheet_name)
        for row_data in sheet_data:
            # Find testpoint columns
            for header in headers:
                if "Testpoint" in header:
                    # Extract testpoint ID
                    parts = header.split("_")
                    for part in parts:
                        if "Testpoint" in part:
                            testpoint_id = part.replace("-", "-")
                            break
                    else:
                        continue
                    
                    if testpoint_id not in testpoint_data:
                        testpoint_data[testpoint_id] = {
                            "coordinates": TESTPOINT_COORDINATES.get(testpoint_id, {"lat": 22.418, "lng": 114.206}),
                            "device": DEVICE_TYPES.get(testpoint_id, {"type": "Unknown", "sensors": [], "category": "unknown"}),
                            "measurements": {}
                        }
                    
                    # Add measurement data
                    if param_name not in testpoint_data[testpoint_id]["measurements"]:
                        testpoint_data[testpoint_id]["measurements"][param_name] = []
                    
                    # Create timestamp from No, Date, Time columns if available
                    timestamp = None
                    if "Date" in row_data and "Time" in row_data:
                        try:
                            date_val = row_data.get("Date", "")
                            time_val = row_data.get("Time", "")
                            if date_val and time_val:
                                if isinstance(date_val, str) and isinstance(time_val, str):
                                    timestamp = f"{date_val} {time_val}"
                                elif hasattr(date_val, 'isoformat'):
                                    timestamp = date_val.isoformat()
                        except:
                            pass
                    
                    value = row_data.get(header)
                    if value is not None and value != "":
                        testpoint_data[testpoint_id]["measurements"][param_name].append({
                            "timestamp": timestamp or row_data.get("No", ""),
                            "value": float(value) if isinstance(value, (int, float)) else value
                        })
    
    return all_data, testpoint_data

def generate_summary(testpoint_data):
    """Generate summary statistics for each testpoint."""
    summary = {}
    
    for testpoint_id, data in testpoint_data.items():
        summary[testpoint_id] = {
            "coordinates": data["coordinates"],
            "device": data["device"],
            "statistics": {}
        }
        
        for param, measurements in data.get("measurements", {}).items():
            values = [m["value"] for m in measurements if isinstance(m.get("value"), (int, float))]
            if values:
                summary[testpoint_id]["statistics"][param] = {
                    "min": round(min(values), 2),
                    "max": round(max(values), 2),
                    "avg": round(sum(values) / len(values), 2),
                    "count": len(values)
                }
    
    return summary

def main():
    # Paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    excel_path = os.path.join(project_root, "public", "Weather data-Monitored.xlsx")
    output_dir = os.path.join(project_root, "public", "data")
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Parse Excel
    all_data, testpoint_data = parse_excel_file(excel_path)
    
    # Generate summary
    summary = generate_summary(testpoint_data)
    
    # Save all data (full)
    with open(os.path.join(output_dir, "monitored_data_full.json"), "w") as f:
        json.dump(all_data, f, indent=2, default=str)
    print(f"Saved full data to {output_dir}/monitored_data_full.json")
    
    # Save testpoint organized data
    with open(os.path.join(output_dir, "testpoint_data.json"), "w") as f:
        json.dump(testpoint_data, f, indent=2, default=str)
    print(f"Saved testpoint data to {output_dir}/testpoint_data.json")
    
    # Save summary
    with open(os.path.join(output_dir, "testpoint_summary.json"), "w") as f:
        json.dump(summary, f, indent=2)
    print(f"Saved summary to {output_dir}/testpoint_summary.json")
    
    # Print summary
    print("\n" + "="*60)
    print("DATA SUMMARY")
    print("="*60)
    for testpoint_id, data in summary.items():
        print(f"\n{testpoint_id}:")
        print(f"  Location: {data['coordinates']['lat']:.6f}°N, {data['coordinates']['lng']:.6f}°E")
        print(f"  Device: {data['device']['type']}")
        if data.get('statistics'):
            print("  Measurements:")
            for param, stats in data['statistics'].items():
                print(f"    {param}: avg={stats['avg']}, range=[{stats['min']}, {stats['max']}], n={stats['count']}")

if __name__ == "__main__":
    main()

