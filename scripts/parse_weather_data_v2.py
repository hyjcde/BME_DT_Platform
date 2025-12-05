#!/usr/bin/env python3
"""
Parse Weather data-Monitored.xlsx and create properly structured JSON data for the thermal DT system.
Version 2: Better organization by test point number with proper coordinates.
"""

import json
import os
import re
from datetime import datetime, time
import openpyxl

# Test point coordinates (adjusted to fit the thermal map - shifted south ~0.003 deg)
# Original Google Maps coordinates were offset from the thermal map base layer
# Map center is approximately 22.4167, 114.2069
TESTPOINT_COORDINATES = {
    1: {"lat": 22.4160, "lng": 114.2077, "name": "Near University Station", "type": "HOBO MX"},
    2: {"lat": 22.4165, "lng": 114.2083, "name": "East Campus Road", "type": "HOBO MX"},
    3: {"lat": 22.4169, "lng": 114.2068, "name": "Central Avenue", "type": "HOBO MX"},
    4: {"lat": 22.4172, "lng": 114.2032, "name": "Northwest Campus", "type": "HOBO MX"},
    5: {"lat": 22.4161, "lng": 114.2047, "name": "West Side", "type": "HOBO MX"},
    6: {"lat": 22.4155, "lng": 114.2044, "name": "Southwest Area", "type": "HOBO MX"},
    7: {"lat": 22.4156, "lng": 114.2056, "name": "Central South", "type": "HOBO MX"},
    8: {"lat": 22.4158, "lng": 114.2065, "name": "Campus Center", "type": "HOBO MX"},
    9: {"lat": 22.4160, "lng": 114.2071, "name": "Weather Station 1", "type": "Weather Station"},
    10: {"lat": 22.4167, "lng": 114.2054, "name": "Weather Station 2", "type": "Weather Station"},
    11: {"lat": 22.4167, "lng": 114.2054, "name": "Thermocouple 1", "type": "Thermocouple"},
    12: {"lat": 22.4160, "lng": 114.2071, "name": "Thermocouple 2", "type": "Thermocouple"},
    13: {"lat": 22.4160, "lng": 114.2071, "name": "Radiation Tracker", "type": "Radiation Tracker"},
}

# Device info
DEVICE_INFO = {
    "HOBO MX": {"sensors": ["Temperature", "RH", "Dew Point", "Light"], "color": "#3b82f6"},
    "Weather Station": {"sensors": ["Temp", "RH", "Globe Temp", "Wind Speed", "Wind Dir", "Solar Rad", "PM2.5", "PM10"], "color": "#22c55e"},
    "Thermocouple": {"sensors": ["Surface Temperature"], "color": "#f59e0b"},
    "Radiation Tracker": {"sensors": ["DHI", "DNI", "GHI"], "color": "#a855f7"},
}

def extract_testpoint_number(header):
    """Extract testpoint number from column header."""
    match = re.search(r'Testpoint[-_]?(\d+)', header, re.IGNORECASE)
    if match:
        return int(match.group(1))
    return None

def parse_time_to_minutes(time_val):
    """Convert time to minutes since midnight for indexing."""
    if isinstance(time_val, time):
        return time_val.hour * 60 + time_val.minute
    if isinstance(time_val, datetime):
        return time_val.hour * 60 + time_val.minute
    if isinstance(time_val, str):
        try:
            t = datetime.strptime(time_val, "%H:%M")
            return t.hour * 60 + t.minute
        except:
            return None
    return None

def normalize_timestamp(date_val, time_val):
    """Normalize date/time values to a consistent timestamp string.
    
    The Time column in Excel contains full datetime (e.g., 2025-10-18 09:09:50).
    We extract both date and time from the time_val if it's a datetime object.
    Returns format: "10-18 09:09" for display.
    """
    # If time_val is a full datetime, extract both date and time from it
    if isinstance(time_val, datetime):
        return time_val.strftime("%m-%d %H:%M")
    
    # Fallback: try to construct from separate date and time
    time_str = None
    date_str = None
    
    if isinstance(time_val, time):
        time_str = time_val.strftime("%H:%M")
    elif isinstance(time_val, str) and ':' in time_val:
        try:
            parts = time_val.split(':')
            time_str = f"{int(parts[0]):02d}:{int(parts[1]):02d}"
        except:
            pass
    
    if isinstance(date_val, datetime):
        date_str = date_val.strftime("%m-%d")
    elif isinstance(date_val, (int, float)):
        from datetime import timedelta
        excel_epoch = datetime(1899, 12, 30)
        actual_date = excel_epoch + timedelta(days=int(date_val))
        date_str = actual_date.strftime("%m-%d")
    
    if date_str and time_str:
        return f"{date_str} {time_str}"
    elif time_str:
        return time_str
    return None

def parse_excel_file(filepath):
    """Parse the Excel file and organize data by testpoint number."""
    print(f"Loading Excel file: {filepath}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Initialize structure for each testpoint
    testpoints = {}
    for tp_num, info in TESTPOINT_COORDINATES.items():
        testpoints[tp_num] = {
            "id": tp_num,
            "name": f"Testpoint-{tp_num}",
            "location_name": info["name"],
            "lat": info["lat"],
            "lng": info["lng"],
            "device_type": info["type"],
            "device_info": DEVICE_INFO.get(info["type"], {}),
            "measurements": {},
            "timeseries": []
        }
    
    # Process each sheet
    sheet_param_mapping = {
        "HOBO_Temp": {"param": "temperature", "unit": "°C"},
        "HOBO_RH": {"param": "relative_humidity", "unit": "%"},
        "HOBO_Light": {"param": "light", "unit": "lux"},
        "HOBO_Dew_point": {"param": "dew_point", "unit": "°C"},
        "Thermocouple_Temp": {"param": "surface_temperature", "unit": "°C"},
        "GlobE_temp": {"param": "globe_temperature", "unit": "°C"},
        "Wind_direction": {"param": "wind_direction", "unit": "°"},
        "Wind_speed": {"param": "wind_speed", "unit": "m/s"},
        "Solar_radiation": {"param": "solar_radiation", "unit": "W/m²"},
        "Air_temp": {"param": "air_temperature", "unit": "°C"},
        "RH": {"param": "rh_station", "unit": "%"},
        "PM10": {"param": "pm10", "unit": "μg/m³"},
        "PM25": {"param": "pm25", "unit": "μg/m³"},
        "Pressure": {"param": "pressure", "unit": "kPa"},
        "Radiation": {"param": "radiation", "unit": "W/m²"},
    }
    
    all_timeseries = {}  # {(date, time): {tp_num: {param: value}}}
    
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        sheet = wb[sheet_name]
        
        param_info = sheet_param_mapping.get(sheet_name, {"param": sheet_name.lower(), "unit": ""})
        param_name = param_info["param"]
        param_unit = param_info["unit"]
        
        # Get headers
        headers = []
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=1, column=col).value
            headers.append(str(val) if val else "")
        
        # Find testpoint columns
        tp_columns = {}  # col_index -> testpoint_number
        for col_idx, header in enumerate(headers):
            tp_num = extract_testpoint_number(header)
            if tp_num:
                tp_columns[col_idx] = tp_num
        
        # Parse data rows
        for row in range(2, min(sheet.max_row + 1, 3000)):  # Limit rows for performance
            # Get date and time
            date_val = None
            time_val = None
            
            for col_idx, header in enumerate(headers):
                if header.lower() == "date":
                    date_val = sheet.cell(row=row, column=col_idx + 1).value
                elif header.lower() == "time":
                    time_val = sheet.cell(row=row, column=col_idx + 1).value
            
            # Get values for each testpoint
            for col_idx, tp_num in tp_columns.items():
                value = sheet.cell(row=row, column=col_idx + 1).value
                if value is not None and isinstance(value, (int, float)):
                    # Store in measurements
                    if param_name not in testpoints[tp_num]["measurements"]:
                        testpoints[tp_num]["measurements"][param_name] = {
                            "unit": param_unit,
                            "values": [],
                            "min": float('inf'),
                            "max": float('-inf'),
                            "sum": 0,
                            "count": 0
                        }
                    
                    m = testpoints[tp_num]["measurements"][param_name]
                    m["values"].append(value)
                    m["min"] = min(m["min"], value)
                    m["max"] = max(m["max"], value)
                    m["sum"] += value
                    m["count"] += 1
                    
                    # Build timeseries key using normalized timestamp
                    ts_key = normalize_timestamp(date_val, time_val)
                    if ts_key:
                        if ts_key not in all_timeseries:
                            all_timeseries[ts_key] = {}
                        if tp_num not in all_timeseries[ts_key]:
                            all_timeseries[ts_key][tp_num] = {}
                        all_timeseries[ts_key][tp_num][param_name] = value
    
    # Finalize measurements
    for tp_num in testpoints:
        for param_name, m in testpoints[tp_num]["measurements"].items():
            if m["count"] > 0:
                m["avg"] = round(m["sum"] / m["count"], 2)
                m["min"] = round(m["min"], 2)
                m["max"] = round(m["max"], 2)
            else:
                m["avg"] = None
                m["min"] = None
                m["max"] = None
            # Remove large values array for summary
            del m["values"]
            del m["sum"]
    
    return testpoints, all_timeseries

def create_animation_data(all_timeseries, testpoints):
    """Create simplified animation data with timestamps."""
    animation_data = []
    
    # Sort timestamps
    sorted_times = sorted(all_timeseries.keys())
    
    for ts in sorted_times:
        frame = {
            "timestamp": ts,
            "testpoints": {}
        }
        
        for tp_num, data in all_timeseries[ts].items():
            frame["testpoints"][tp_num] = {
                "temperature": data.get("temperature") or data.get("air_temperature") or data.get("surface_temperature"),
                "humidity": data.get("relative_humidity") or data.get("rh_station"),
                "wind_speed": data.get("wind_speed"),
                "solar_radiation": data.get("solar_radiation"),
            }
        
        animation_data.append(frame)
    
    return animation_data

def create_summary(testpoints):
    """Create a clean summary for the frontend."""
    summary = []
    
    for tp_num, data in sorted(testpoints.items()):
        tp_summary = {
            "id": tp_num,
            "name": data["name"],
            "location_name": data["location_name"],
            "lat": data["lat"],
            "lng": data["lng"],
            "device_type": data["device_type"],
            "color": data["device_info"].get("color", "#888"),
            "current_values": {},
            "statistics": {}
        }
        
        # Add measurement statistics
        for param, stats in data["measurements"].items():
            tp_summary["statistics"][param] = {
                "min": stats["min"],
                "max": stats["max"],
                "avg": stats["avg"],
                "unit": stats["unit"],
                "count": stats["count"]
            }
            
            # Set "current" as average for demo
            if stats["avg"]:
                tp_summary["current_values"][param] = stats["avg"]
        
        summary.append(tp_summary)
    
    return summary

def main():
    # Paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    excel_path = os.path.join(project_root, "public", "Weather data-Monitored.xlsx")
    output_dir = os.path.join(project_root, "public", "data")
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Parse Excel
    testpoints, all_timeseries = parse_excel_file(excel_path)
    
    # Create summary
    summary = create_summary(testpoints)
    
    # Create animation data (sampled for performance)
    animation_data = create_animation_data(all_timeseries, testpoints)
    
    # Sample animation data (every 10 minutes = every 10 frames approximately)
    animation_sampled = animation_data[::10] if len(animation_data) > 100 else animation_data
    
    # Save summary
    with open(os.path.join(output_dir, "testpoints.json"), "w") as f:
        json.dump(summary, f, indent=2)
    print(f"Saved testpoints summary to {output_dir}/testpoints.json")
    
    # Save animation data
    with open(os.path.join(output_dir, "timeseries.json"), "w") as f:
        json.dump(animation_sampled, f, indent=2)
    print(f"Saved timeseries data to {output_dir}/timeseries.json ({len(animation_sampled)} frames)")
    
    # Print summary
    print("\n" + "="*70)
    print("TESTPOINT SUMMARY")
    print("="*70)
    for tp in summary:
        print(f"\n[{tp['id']:2d}] {tp['name']} - {tp['location_name']}")
        print(f"     Location: {tp['lat']:.6f}°N, {tp['lng']:.6f}°E")
        print(f"     Device: {tp['device_type']}")
        if tp['statistics']:
            print("     Data:")
            for param, stats in tp['statistics'].items():
                if stats['avg']:
                    print(f"       {param}: {stats['avg']:.1f} [{stats['min']:.1f}-{stats['max']:.1f}] {stats['unit']} (n={stats['count']})")

if __name__ == "__main__":
    main()

