#!/usr/bin/env python3
"""
Generate monitored-data JSON artifacts from the field Excel workbook.
"""

from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from copy import deepcopy
from dataclasses import dataclass
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel


PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_ROOT / "Weather data monitored_Oct_18-19_2025.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "public" / "data"


TESTPOINT_METADATA: dict[int, dict[str, Any]] = {
    1: {"name": "Testpoint-1", "location_name": "Near University Station", "lat": 22.419000, "lng": 114.207738, "device_type": "HOBO MX", "color": "#3b82f6"},
    2: {"name": "Testpoint-2", "location_name": "East Campus Road", "lat": 22.419548, "lng": 114.208326, "device_type": "HOBO MX", "color": "#3b82f6"},
    3: {"name": "Testpoint-3", "location_name": "Central Avenue", "lat": 22.419937, "lng": 114.206832, "device_type": "HOBO MX", "color": "#3b82f6"},
    4: {"name": "Testpoint-4", "location_name": "Northwest Campus", "lat": 22.420221, "lng": 114.203237, "device_type": "HOBO MX", "color": "#3b82f6"},
    5: {"name": "Testpoint-5", "location_name": "West Side", "lat": 22.419147, "lng": 114.204707, "device_type": "HOBO MX", "color": "#3b82f6"},
    6: {"name": "Testpoint-6", "location_name": "Southwest Area", "lat": 22.418473, "lng": 114.204404, "device_type": "HOBO MX", "color": "#3b82f6"},
    7: {"name": "Testpoint-7", "location_name": "Central South", "lat": 22.418608, "lng": 114.205645, "device_type": "HOBO MX", "color": "#3b82f6"},
    8: {"name": "Testpoint-8", "location_name": "Campus Center", "lat": 22.418800, "lng": 114.206500, "device_type": "HOBO MX", "color": "#3b82f6"},
    9: {"name": "Testpoint-9", "location_name": "Weather Station 1", "lat": 22.418964, "lng": 114.207135, "device_type": "Weather Station", "color": "#22c55e"},
    10: {"name": "Testpoint-10", "location_name": "Weather Station 2", "lat": 22.419745, "lng": 114.205381, "device_type": "Weather Station", "color": "#22c55e"},
    11: {"name": "Testpoint-11", "location_name": "Thermocouple 1", "lat": 22.419745, "lng": 114.205381, "device_type": "Thermocouple", "color": "#f59e0b"},
    12: {"name": "Testpoint-12", "location_name": "Thermocouple 2", "lat": 22.418964, "lng": 114.207135, "device_type": "Thermocouple", "color": "#f59e0b"},
    13: {"name": "Testpoint-13", "location_name": "Radiation Tracker", "lat": 22.418964, "lng": 114.207135, "device_type": "Radiation Tracker", "color": "#a855f7"},
}


METRIC_UNITS: dict[str, str] = {
    "temperature": "°C",
    "humidity": "%",
    "light": "lux",
    "dew_point": "°C",
    "air_temperature": "°C",
    "globe_temperature": "°C",
    "surface_temperature": "°C",
    "wind_speed": "m/s",
    "wind_direction": "°",
    "solar_radiation": "W/m²",
    "pressure": "kPa",
    "pm10": "μg/m³",
    "pm25": "μg/m³",
    "diffuse_radiation": "W/m²",
    "direct_normal_radiation": "W/m²",
    "direct_horizontal_radiation": "W/m²",
}


SHEET_METRIC_MAP: dict[str, str] = {
    "HOBO_Temp": "temperature",
    "HOBO_RH": "humidity",
    "HOBO_Light": "light",
    "HOBO_Dew_point": "dew_point",
    "GlobE_temp": "globe_temperature",
    "Wind_direction": "wind_direction",
    "Wind_speed": "wind_speed",
    "Solar_radiation": "solar_radiation",
    "Air_temp": "air_temperature",
    "RH": "humidity",
    "PM10": "pm10",
    "PM25": "pm25",
    "Pressure": "pressure",
}


RADIATION_HEADER_MAP: dict[str, str] = {
    "Diffuse radiation": "diffuse_radiation",
    "Normal direct radiation": "direct_normal_radiation",
    "Horizontal direct radiation": "direct_horizontal_radiation",
    "Solar_radiation": "solar_radiation",
}


@dataclass
class MeasurementRecord:
    timestamp: str
    value: float


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and not math.isnan(float(value))


def normalize_value(value: Any) -> float | None:
    if not is_number(value):
        return None
    return round(float(value), 6)


def extract_testpoint_id(header: Any) -> int | None:
    if header is None:
        return None
    match = re.search(r"Test\w*-?(\d+)", str(header), re.IGNORECASE)
    if not match:
        return None
    return int(match.group(1))


def excel_date_to_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        converted = from_excel(value)
        if isinstance(converted, datetime):
            return converted.date()
        if isinstance(converted, time):
            return None
    return None


def build_timestamp(date_value: Any, time_value: Any) -> str | None:
    if isinstance(time_value, datetime):
        return time_value.replace(microsecond=0).isoformat()

    base_date = excel_date_to_date(date_value)
    if isinstance(time_value, time) and base_date:
        return datetime.combine(base_date, time_value).replace(microsecond=0).isoformat()

    if isinstance(time_value, (int, float)):
        converted = from_excel(time_value)
        if isinstance(converted, datetime):
            return converted.replace(microsecond=0).isoformat()
        if isinstance(converted, time) and base_date:
            return datetime.combine(base_date, converted).replace(microsecond=0).isoformat()

    if base_date and isinstance(time_value, str):
        for fmt in ("%H:%M:%S", "%H:%M"):
            try:
                parsed = datetime.strptime(time_value, fmt).time()
                return datetime.combine(base_date, parsed).replace(microsecond=0).isoformat()
            except ValueError:
                continue

    return None


def parse_sheet_header(header_row: tuple[Any, ...]) -> list[str]:
    return [str(cell).strip() if cell is not None else "" for cell in header_row]


def process_standard_sheet(
    sheet_name: str,
    headers: list[str],
    rows: list[tuple[Any, ...]],
    per_point_measurements: dict[int, dict[str, list[MeasurementRecord]]],
    frames: dict[str, dict[int, dict[str, float | None]]],
) -> None:
    metric_key = SHEET_METRIC_MAP[sheet_name]
    date_idx = headers.index("Date")
    time_idx = headers.index("Time")
    point_columns = {idx: extract_testpoint_id(header) for idx, header in enumerate(headers)}

    for row in rows:
        timestamp = build_timestamp(row[date_idx], row[time_idx])
        if not timestamp:
            continue

        for column_idx, point_id in point_columns.items():
            if not point_id:
                continue
            value = normalize_value(row[column_idx])
            if value is None:
                continue

            per_point_measurements[point_id][metric_key].append(MeasurementRecord(timestamp=timestamp, value=value))
            frames[timestamp][point_id][metric_key] = value


def process_thermocouple_sheet(
    headers: list[str],
    rows: list[tuple[Any, ...]],
    per_point_measurements: dict[int, dict[str, list[MeasurementRecord]]],
    frames: dict[str, dict[int, dict[str, float | None]]],
) -> None:
    date_idx = headers.index("Date")
    time_idx = headers.index("Time")

    target_columns: dict[int, int] = {}
    for idx, header in enumerate(headers):
        if "LST_Testpoint-" not in header:
            continue
        point_id = extract_testpoint_id(header)
        if point_id:
            target_columns[idx] = point_id

    for row in rows:
        timestamp = build_timestamp(row[date_idx], row[time_idx])
        if not timestamp:
            continue

        for column_idx, point_id in target_columns.items():
            value = normalize_value(row[column_idx])
            if value is None:
                continue
            per_point_measurements[point_id]["surface_temperature"].append(MeasurementRecord(timestamp=timestamp, value=value))
            frames[timestamp][point_id]["surface_temperature"] = value


def process_radiation_sheet(
    headers: list[str],
    rows: list[tuple[Any, ...]],
    per_point_measurements: dict[int, dict[str, list[MeasurementRecord]]],
    frames: dict[str, dict[int, dict[str, float | None]]],
) -> None:
    date_idx = headers.index("Date")
    time_idx = headers.index("Time")
    point_columns: dict[int, tuple[int, str]] = {}

    for idx, header in enumerate(headers):
        point_id = extract_testpoint_id(header)
        if not point_id:
            continue
        for fragment, metric_key in RADIATION_HEADER_MAP.items():
            if fragment.lower() in header.lower():
                point_columns[idx] = (point_id, metric_key)
                break

    for row in rows:
        timestamp = build_timestamp(row[date_idx], row[time_idx])
        if not timestamp:
            continue

        for column_idx, (point_id, metric_key) in point_columns.items():
            value = normalize_value(row[column_idx])
            if value is None:
                continue
            per_point_measurements[point_id][metric_key].append(MeasurementRecord(timestamp=timestamp, value=value))
            frames[timestamp][point_id][metric_key] = value


def collect_workbook_data() -> tuple[
    dict[int, dict[str, list[MeasurementRecord]]],
    dict[str, dict[int, dict[str, float | None]]],
    dict[str, dict[str, Any]],
]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    per_point_measurements: dict[int, dict[str, list[MeasurementRecord]]] = defaultdict(lambda: defaultdict(list))
    frames: dict[str, dict[int, dict[str, float | None]]] = defaultdict(lambda: defaultdict(dict))
    workbook_dump: dict[str, dict[str, Any]] = {}

    for sheet in workbook.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter)
        headers = parse_sheet_header(header_row)
        rows = [row for row in rows_iter if any(cell is not None for cell in row)]
        workbook_dump[sheet.title] = {
            "headers": headers,
            "data": [
                {
                    headers[idx]: (
                        cell.replace(microsecond=0).isoformat()
                        if isinstance(cell, datetime)
                        else cell.isoformat() if isinstance(cell, time)
                        else cell
                    )
                    for idx, cell in enumerate(row)
                    if headers[idx]
                }
                for row in rows
            ],
        }

        if sheet.title in SHEET_METRIC_MAP:
            process_standard_sheet(sheet.title, headers, rows, per_point_measurements, frames)
        elif sheet.title == "Thermocouple_Temp":
            process_thermocouple_sheet(headers, rows, per_point_measurements, frames)
        elif sheet.title == "Radiation":
            process_radiation_sheet(headers, rows, per_point_measurements, frames)

    return per_point_measurements, frames, workbook_dump


def resolve_metric_source(point_id: int, metric_key: str) -> str:
    if metric_key == "humidity":
        return "HOBO_RH" if point_id <= 7 else "RH"
    if metric_key == "solar_radiation":
        return "Radiation" if point_id == 13 else "Solar_radiation"
    if metric_key in {"diffuse_radiation", "direct_normal_radiation", "direct_horizontal_radiation"}:
        return "Radiation"
    if metric_key == "surface_temperature":
        return "Thermocouple_Temp"
    return {
        "temperature": "HOBO_Temp",
        "light": "HOBO_Light",
        "dew_point": "HOBO_Dew_point",
        "air_temperature": "Air_temp",
        "globe_temperature": "GlobE_temp",
        "wind_speed": "Wind_speed",
        "wind_direction": "Wind_direction",
        "pressure": "Pressure",
        "pm10": "PM10",
        "pm25": "PM25",
    }.get(metric_key, "")


def build_statistics(
    point_id: int,
    point_measurements: dict[str, list[MeasurementRecord]],
) -> tuple[dict[str, Any], dict[str, float]]:
    statistics: dict[str, Any] = {}
    current_values: dict[str, float] = {}

    for metric_key, records in sorted(point_measurements.items()):
        if not records:
            continue
        values = [record.value for record in records]
        statistics[metric_key] = {
            "min": round(min(values), 2),
            "max": round(max(values), 2),
            "avg": round(sum(values) / len(values), 2),
            "unit": METRIC_UNITS.get(metric_key, ""),
            "count": len(values),
            "source_sheet": resolve_metric_source(point_id, metric_key),
        }
        current_values[metric_key] = round(records[-1].value, 2)

    return statistics, current_values


def build_testpoints_json(
    per_point_measurements: dict[int, dict[str, list[MeasurementRecord]]],
) -> list[dict[str, Any]]:
    testpoints: list[dict[str, Any]] = []

    for point_id in sorted(TESTPOINT_METADATA):
        metadata = deepcopy(TESTPOINT_METADATA[point_id])
        measurements = per_point_measurements.get(point_id, {})
        statistics, current_values = build_statistics(point_id, measurements)
        testpoints.append(
            {
                "id": point_id,
                **metadata,
                "current_values": current_values,
                "statistics": statistics,
                "available_metrics": sorted(statistics.keys()),
            }
        )

    return testpoints


def build_timeseries_json(frames: dict[str, dict[int, dict[str, float | None]]]) -> list[dict[str, Any]]:
    timeseries: list[dict[str, Any]] = []
    for timestamp in sorted(frames):
        frame_points = {
            str(point_id): dict(sorted(metrics.items()))
            for point_id, metrics in sorted(frames[timestamp].items())
        }
        timeseries.append({"timestamp": timestamp, "testpoints": frame_points})
    return timeseries


def build_testpoint_data_json(
    per_point_measurements: dict[int, dict[str, list[MeasurementRecord]]],
) -> dict[str, Any]:
    output: dict[str, Any] = {}
    for point_id in sorted(TESTPOINT_METADATA):
        metadata = TESTPOINT_METADATA[point_id]
        measurements = per_point_measurements.get(point_id, {})
        output[str(point_id)] = {
            "metadata": {"id": point_id, **metadata},
            "measurements": {
                metric_key: [
                    {"timestamp": record.timestamp, "value": record.value}
                    for record in records
                ]
                for metric_key, records in sorted(measurements.items())
            },
        }
    return output


def build_testpoint_summary_json(testpoints: list[dict[str, Any]], timeseries: list[dict[str, Any]]) -> dict[str, Any]:
    metric_catalog = sorted({metric for point in testpoints for metric in point["statistics"].keys()})
    return {
        "generated_at": datetime.now(UTC).replace(microsecond=0).isoformat(),
        "source_file": WORKBOOK_PATH.name,
        "measurement_period": {
            "start": timeseries[0]["timestamp"] if timeseries else None,
            "end": timeseries[-1]["timestamp"] if timeseries else None,
            "frame_count": len(timeseries),
        },
        "metric_catalog": metric_catalog,
        "testpoints": [
            {
                "id": point["id"],
                "name": point["name"],
                "location_name": point["location_name"],
                "device_type": point["device_type"],
                "available_metrics": point["available_metrics"],
            }
            for point in testpoints
        ],
    }


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    per_point_measurements, frames, workbook_dump = collect_workbook_data()
    testpoints = build_testpoints_json(per_point_measurements)
    timeseries = build_timeseries_json(frames)
    testpoint_data = build_testpoint_data_json(per_point_measurements)
    testpoint_summary = build_testpoint_summary_json(testpoints, timeseries)

    write_json(OUTPUT_DIR / "testpoints.json", testpoints)
    write_json(OUTPUT_DIR / "timeseries.json", timeseries)
    write_json(OUTPUT_DIR / "testpoint_data.json", testpoint_data)
    write_json(OUTPUT_DIR / "testpoint_summary.json", testpoint_summary)
    write_json(OUTPUT_DIR / "monitored_data_full.json", workbook_dump)

    print(f"Generated testpoints: {len(testpoints)}")
    print(f"Generated timeseries frames: {len(timeseries)}")
    print(f"Output written to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
